from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 복사할 파일 불러오기
source_wb = load_workbook("거래내역.xlsx")
source_ws = source_wb.active

# 붙여넣을 파일 불러오기
destination_wb = load_workbook('대체투자요약_updated(수식적용).xlsx')
destination_ws = destination_wb["거래내역"]

# 복사할 범위 선택 (전체 범위)
copy_range = source_ws.dimensions



# 선택한 범위 데이터 복사
copied_data = []
for row in source_ws[copy_range]:
    row_data = []
    for cell in row:
        value = cell.value
        if isinstance(value, str) :
            # 숫자에 포함된 쉼표 제거
            value = value.replace(',', '')
        row_data.append(value)
    copied_data.append(row_data)

# 붙여넣기할 위치 찾기
last_row = destination_ws.max_row + 1

#선택한 범위 데이터 붙여넣기
for row_data in copied_data:
    destination_ws.append(row_data)




# 변경사항 저장
destination_wb.save('대체투자요약_updated(거래반영).xlsx')