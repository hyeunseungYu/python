import urllib.parse
import requests
import time
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import os
from collections import Counter
from tqdm import tqdm



def search_naver_news(query):
    encoded_query = urllib.parse.quote(query)

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36"
    }

    news_data = []

    pbar = tqdm(total=3, desc='page')

    for page in range(1, 4):
        url = f"https://search.naver.com/search.naver?where=news&sm=tab_jum&query={encoded_query}&sort=1&photo=0&field=0&reporter_article=&pd=3&ds=&de=&docid=&nso=so:r,p:from20210501to20210510,a:all&mynews=0&cluster_rank=52&start={(page - 1) * 10 + 1}&refresh_start=0"

        response = requests.get(url, headers=headers, timeout=5)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, "html.parser")

        
        news_articles = soup.select("#main_pack > section > div > div.group_news > ul > li")


        for article in news_articles:
            source = article.select_one('.info.press').text
            title = article.select_one('.news_tit').text
            link = article.select_one('.news_tit')['href']
            
            news_data.append({'Source':source,'Title':title, 'Link':link})
            
            time.sleep(0.5)
        
        pbar.set_description(f'{page}번 페이지 완료')
        pbar.update(1)


    
    df = pd.DataFrame(news_data, columns=["Source", "Title", "Link"])
    
    
    # 제목에서 단어 추출
    titles = df["Title"].str.lower().str.split()
    words = [word for sublist in titles for word in sublist]
    
    # 가장 많이 나타나는 단어 찾기
    word_counts = Counter(words)
    most_common_word = word_counts.most_common(1)[0][0]
    
    print(f"가장 많이 나타나는 단어: {most_common_word}")

    # 파일 저장 경로 설정
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    folder_path = os.path.join(desktop_path, "기사검색")
    os.makedirs(folder_path, exist_ok=True)
    file_name = f"{query}_{time.strftime('%Y%m%d')}.xlsx"
    file_path = os.path.join(folder_path, file_name)

    # 엑셀 파일 생성
    workbook = Workbook()
    worksheet = workbook.active

    # 데이터 프레임을 엑셀에 추가
    for row in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(row)

     # 가장 많이 나타나는 단어를 노란색으로 채우기
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in worksheet["B"]:
        if cell.value == most_common_word:
            cell.fill = fill

    # 엑셀 파일 저장
    workbook.save(file_path)
    print(f"파일 다운로드 완료: {file_path}")

# 기사 검색어 입력
query = input("기사 검색어를 입력하세요: ")
search_naver_news(query)